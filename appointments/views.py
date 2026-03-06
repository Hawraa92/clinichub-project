# appointments/views.py
from __future__ import annotations

import base64
import csv
import io
import json
import secrets
from datetime import date, timedelta
from functools import wraps
from typing import Any, Callable
from urllib.parse import urlencode

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDate
from django.http import HttpRequest, HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template import TemplateDoesNotExist
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.timezone import get_default_timezone, make_aware
from django.views.decorators.cache import cache_control
from django.views.decorators.http import require_GET, require_POST, require_http_methods

from accounts.forms import CustomPasswordForm, ProfileUpdateForm
from doctor.models import Doctor
from patient.forms import SecretaryPatientForm
from patient.models import Patient

from .forms import AppointmentForm, DATETIME_INPUT_FORMATS, DateTimeLocalInput
from .models import Appointment, AppointmentStatus, Notification, PatientBookingRequest

# ✅ Audit Log (optional-safe import)
try:  # pragma: no cover
    from audit.utils import log_action as _audit_log_action
except Exception:  # pragma: no cover
    _audit_log_action = None  # type: ignore

# Optional BookingRequestStatus (if your build includes it)
try:  # pragma: no cover
    from .models import BookingRequestStatus  # type: ignore
except Exception:  # pragma: no cover
    BookingRequestStatus = None  # type: ignore

User = get_user_model()

# ------------------------------------------------------------------#
#                   Time normalization (robust)                      #
# ------------------------------------------------------------------#
_LOCAL_TZ = get_default_timezone()

# ✅ Prefer using the SAME helpers from models.py (avoid mismatch)
try:  # pragma: no cover
    from .models import _normalize_dt as _model_normalize_dt  # type: ignore
    from .models import _now_local as _model_now_local  # type: ignore
    from .models import _past_tolerance as _model_past_tolerance  # type: ignore
except Exception:  # pragma: no cover

    def _model_now_local():
        now = timezone.now()
        if bool(getattr(settings, "USE_TZ", False)):
            if timezone.is_naive(now):
                now = make_aware(now, _LOCAL_TZ)
            return timezone.localtime(now, _LOCAL_TZ)
        if timezone.is_aware(now):
            now = timezone.localtime(now, _LOCAL_TZ).replace(tzinfo=None)
        return now

    def _model_normalize_dt(dt):
        if dt is None:
            return None
        use_tz = bool(getattr(settings, "USE_TZ", False))
        if use_tz:
            if timezone.is_naive(dt):
                return make_aware(dt, _LOCAL_TZ)
            return timezone.localtime(dt, _LOCAL_TZ)
        if timezone.is_aware(dt):
            return timezone.localtime(dt, _LOCAL_TZ).replace(tzinfo=None)
        return dt

    def _model_past_tolerance():
        return timedelta(seconds=60)


def _normalize_dt(dt):
    return _model_normalize_dt(dt)


def _now_local():
    return _model_now_local()


def _past_tolerance() -> timedelta:
    return _model_past_tolerance()


def _is_past(dt) -> bool:
    nd = _normalize_dt(dt)
    if nd is None:
        return False
    return nd < (_now_local() - _past_tolerance())


def _dt_for_display(dt):
    nd = _normalize_dt(dt)
    if nd is None:
        return None
    if timezone.is_aware(nd):
        return timezone.localtime(nd, _LOCAL_TZ)
    return nd


def _fmt_dt(dt, fmt: str) -> str:
    d = _dt_for_display(dt)
    return d.strftime(fmt) if d else ""


# ------------------------------------------------------------------#
#                           Helpers                                  #
# ------------------------------------------------------------------#
def _json_success(data: dict[str, Any]) -> JsonResponse:
    return JsonResponse({**data, "success": True})


def _json_error(msg: str, *, status: int = 400) -> JsonResponse:
    return JsonResponse({"success": False, "error": msg}, status=status)


def _today() -> date:
    return timezone.localdate()


def _model_has_field(model, field_name: str) -> bool:
    try:
        model._meta.get_field(field_name)  # type: ignore[attr-defined]
        return True
    except Exception:
        return False


def _active_appts_qs():
    """
    Returns active appointments queryset if soft-delete field exists,
    otherwise returns all appointments.
    """
    qs = Appointment.objects.all()
    try:
        return qs.filter(is_deleted=False)
    except Exception:
        return qs


def _doctor_name(doc: Doctor | None) -> str:
    if not doc:
        return "Doctor"
    try:
        display = getattr(doc, "get_display_name", None)
        if callable(display):
            v = (display() or "").strip()
            if v:
                return v
    except Exception:
        pass

    full_name = (getattr(doc, "full_name", "") or "").strip()
    if full_name:
        return full_name

    user = getattr(doc, "user", None)
    if user:
        try:
            v = (user.get_full_name() or "").strip()
            if v:
                return v
        except Exception:
            pass
        first = (getattr(user, "first_name", "") or "").strip()
        if first:
            return first
        username = (getattr(user, "username", "") or "").strip()
        if username:
            return username.split("@")[0] if "@" in username else username

    return "Doctor"


def _user_name(u) -> str:
    if not u:
        return "User"
    try:
        v = (u.get_full_name() or "").strip()
        if v:
            return v
    except Exception:
        pass
    first = (getattr(u, "first_name", "") or "").strip()
    if first:
        return first
    username = (getattr(u, "username", "") or "").strip()
    if username:
        return username.split("@")[0] if "@" in username else username
    return "User"


def _user_is_secretary(user) -> bool:
    role = getattr(user, "role", None)
    if hasattr(User, "Roles"):
        try:
            return role == User.Roles.SECRETARY
        except Exception:
            pass
    return role == "secretary"


def _user_is_patient(user) -> bool:
    role = getattr(user, "role", None)
    if hasattr(User, "Roles"):
        try:
            return role == User.Roles.PATIENT
        except Exception:
            pass
    return role == "patient"


def _audit(
    *,
    request: HttpRequest | None = None,
    actor=None,
    action: str = "other",
    instance=None,
    message: str = "",
    extra_data: dict[str, Any] | None = None,
) -> None:
    """Safe wrapper for audit logging; never breaks flow."""
    if _audit_log_action is None:
        return
    try:
        _audit_log_action(
            request=request,
            actor=actor,
            action=action,
            instance=instance,
            message=message,
            extra_data=extra_data or {},
        )
    except Exception:
        return


def _secretary_assigned_doctor(user) -> Doctor | None:
    """
    Best-effort resolver for a secretary's assigned doctor across different schemas.
    Caches result on the user object for current request lifecycle.
    """
    if getattr(user, "is_superuser", False):
        return None
    if not _user_is_secretary(user):
        return None

    cache_attr = "_cached_assigned_doctor_obj"
    if hasattr(user, cache_attr):
        return getattr(user, cache_attr)

    doc: Doctor | None = None
    try:
        direct = getattr(user, "assigned_doctor", None)
        if isinstance(direct, Doctor):
            doc = direct
        else:
            direct_id = getattr(user, "assigned_doctor_id", None)
            if direct_id:
                doc = Doctor.objects.select_related("user").filter(pk=direct_id).first()

        if doc is None:
            for attr in ("secretary_profile", "secretary", "profile", "staff_profile"):
                obj = getattr(user, attr, None)
                if not obj:
                    continue

                cand = getattr(obj, "doctor", None) or getattr(obj, "assigned_doctor", None)
                if isinstance(cand, Doctor):
                    doc = cand
                    break

                cand_id = getattr(obj, "doctor_id", None) or getattr(obj, "assigned_doctor_id", None)
                if cand_id:
                    doc = Doctor.objects.select_related("user").filter(pk=cand_id).first()
                    break
    except Exception:
        doc = None

    setattr(user, cache_attr, doc)
    return doc


def _filter_appointments_for_user(qs, user):
    if user is None:
        return qs
    if getattr(user, "is_superuser", False):
        return qs

    assigned = _secretary_assigned_doctor(user)
    if assigned is not None:
        return qs.filter(doctor_id=assigned.id)

    if _user_is_secretary(user) and getattr(settings, "STRICT_SECRETARY_SCOPE", False):
        return qs.none()

    return qs


def _filter_booking_requests_for_user(qs, user):
    if user is None:
        return qs
    if getattr(user, "is_superuser", False):
        return qs

    assigned = _secretary_assigned_doctor(user)
    if assigned is not None:
        return qs.filter(doctor_id=assigned.id)

    if _user_is_secretary(user) and getattr(settings, "STRICT_SECRETARY_SCOPE", False):
        return qs.none()

    return qs


def _doctor_specialty(doc: Doctor) -> str:
    for field in ("speciality", "specialty", "specialization", "department"):
        if _model_has_field(Doctor, field):
            return str(getattr(doc, field, "") or "")
    return ""


def _doctor_room(doc: Doctor) -> str:
    for field in ("room_number", "room", "clinic_room"):
        if _model_has_field(Doctor, field):
            return str(getattr(doc, field, "") or "")
    return ""


def _secretary_default_status():
    for name in ("APPROVED", "CONFIRMED", "ACCEPTED"):
        if hasattr(AppointmentStatus, name):
            return getattr(AppointmentStatus, name)
    return AppointmentStatus.PENDING


def _redirect_with_query(
    viewname: str,
    *,
    query: dict[str, Any] | None = None,
    kwargs: dict[str, Any] | None = None,
):
    url = reverse(viewname, kwargs=kwargs or None)
    if query:
        qs = urlencode({k: v for k, v in query.items() if v is not None and v != ""})
        if qs:
            url = f"{url}?{qs}"
    return redirect(url)


def _safe_reverse(name: str, default: str) -> str:
    try:
        return reverse(name)
    except (NoReverseMatch, Exception):
        return default


def _safe_next_url(request: HttpRequest, fallback: str) -> str:
    nxt = (request.POST.get("next") or request.GET.get("next") or "").strip()
    if not nxt:
        return fallback
    if url_has_allowed_host_and_scheme(
        url=nxt,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return nxt
    return fallback


def _extract_validation_message(
    exc: ValidationError,
    *,
    default: str,
    preferred_field: str = "scheduled_time",
) -> str:
    msg = default
    if hasattr(exc, "message_dict") and isinstance(getattr(exc, "message_dict", None), dict):
        field_msgs = exc.message_dict.get(preferred_field)
        if isinstance(field_msgs, (list, tuple)) and field_msgs:
            return f"❌ {field_msgs[0]}"
    if getattr(exc, "messages", None):
        return f"❌ {exc.messages[0]}"
    return msg


def _lock_form_doctor_field(form: forms.Form, assigned_doctor: Doctor | None) -> None:
    if assigned_doctor is None:
        return
    if "doctor" not in form.fields:
        return

    form.fields["doctor"].queryset = Doctor.objects.filter(pk=assigned_doctor.pk)
    form.fields["doctor"].initial = assigned_doctor.pk
    try:
        form.fields["doctor"].empty_label = None
    except Exception:
        pass
    try:
        form.fields["doctor"].disabled = True
    except Exception:
        pass
    try:
        form.fields["doctor"].widget.attrs.update({"disabled": "disabled"})
    except Exception:
        pass


def _mark_booking_request_seen(br: PatientBookingRequest | None) -> None:
    if br is None:
        return
    updates: dict[str, Any] = {}
    if _model_has_field(PatientBookingRequest, "seen_by_secretary"):
        updates["seen_by_secretary"] = True
    if _model_has_field(PatientBookingRequest, "seen_at"):
        updates["seen_at"] = timezone.now()
    if updates:
        PatientBookingRequest.objects.filter(pk=br.pk).update(**updates)


def _mark_booking_requests_seen_bulk(ids: list[int]) -> None:
    if not ids:
        return
    updates: dict[str, Any] = {}
    if _model_has_field(PatientBookingRequest, "seen_by_secretary"):
        updates["seen_by_secretary"] = True
    if _model_has_field(PatientBookingRequest, "seen_at"):
        updates["seen_at"] = timezone.now()
    if updates:
        PatientBookingRequest.objects.filter(pk__in=ids).update(**updates)


def _notif_has_related_request() -> bool:
    return _model_has_field(Notification, "related_booking_request")


def _mark_related_notifications_read(br: PatientBookingRequest | None) -> None:
    if br is None or not _notif_has_related_request():
        return
    try:
        Notification.objects.filter(related_booking_request=br).update(is_read=True)
    except Exception:
        pass


def secretary_required(view: Callable):
    @wraps(view)
    @login_required
    def wrapper(request, *a, **kw):
        if (not _user_is_secretary(request.user)) and (not request.user.is_superuser):
            return HttpResponseForbidden("You do not have permission to access this page.")
        return view(request, *a, **kw)

    return wrapper


def staff_ticket_required(view: Callable):
    @wraps(view)
    @login_required
    def wrapper(request, *a, **kw):
        if request.user.is_superuser:
            return view(request, *a, **kw)
        if _user_is_secretary(request.user):
            return view(request, *a, **kw)
        return HttpResponseForbidden("You do not have permission to access this page.")

    return wrapper


def is_patient(user) -> bool:
    return _user_is_patient(user)


def _logout_other_sessions(request: HttpRequest) -> None:
    current_key = request.session.session_key
    if not current_key:
        return

    user_id = str(request.user.id)
    qs = Session.objects.filter(expire_date__gte=timezone.now()).exclude(session_key=current_key)

    for s in qs:
        try:
            data = s.get_decoded()
            if str(data.get("_auth_user_id")) == user_id:
                s.delete()
        except Exception:
            continue


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (TypeError, ValueError):
        return None


def _get_period_range(request: HttpRequest, default_period: str = "day") -> tuple[str, date, date]:
    today = _today()
    period = (request.GET.get("period") or default_period).lower()
    if period not in {"day", "week", "month", "custom"}:
        period = default_period

    if period == "day":
        start = end = today
    elif period == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif period == "month":
        start = today.replace(day=1)
        if start.month == 12:
            next_month = start.replace(year=start.year + 1, month=1, day=1)
        else:
            next_month = start.replace(month=start.month + 1, day=1)
        end = next_month - timedelta(days=1)
    else:
        s = _parse_iso_date(request.GET.get("start"))
        e = _parse_iso_date(request.GET.get("end"))
        start = s or today
        end = e or start
        if end < start:
            start, end = end, start

    return period, start, end


def _queue_active_statuses() -> list:
    statuses = [AppointmentStatus.PENDING]
    for name in ("APPROVED", "CONFIRMED", "ACCEPTED"):
        if hasattr(AppointmentStatus, name):
            statuses.append(getattr(AppointmentStatus, name))
    if hasattr(AppointmentStatus, "CALLED"):
        statuses.append(getattr(AppointmentStatus, "CALLED"))
    seen = set()
    out = []
    for s in statuses:
        if s not in seen:
            out.append(s)
            seen.add(s)
    return out


def _queue_waiting_statuses() -> list:
    active = _queue_active_statuses()
    called = getattr(AppointmentStatus, "CALLED", None)
    if called is None:
        return active
    return [s for s in active if s != called]


def _queue_use_called_status() -> bool:
    return bool(getattr(settings, "QUEUE_USE_CALLED_STATUS", False)) and hasattr(AppointmentStatus, "CALLED")


def _get_patient_for_user(user) -> Patient | None:
    p = getattr(user, "patient_profile", None) or getattr(user, "patient", None)
    if isinstance(p, Patient):
        return p
    if _model_has_field(Patient, "user"):
        return Patient.objects.filter(user=user).first()
    return None


def _format_queue_number(n: int | None) -> str:
    return f"P-{int(n):03d}" if n else "-"


def _filter_by_day(qs, day_val: date):
    if _model_has_field(Appointment, "scheduled_day"):
        return qs.filter(scheduled_day=day_val)
    return qs.filter(scheduled_time__date=day_val)


def _filter_by_day_range(qs, start: date, end: date):
    if _model_has_field(Appointment, "scheduled_day"):
        return qs.filter(scheduled_day__range=(start, end))
    return qs.filter(scheduled_time__date__range=(start, end))


def _set_booking_request_status(br: PatientBookingRequest, target_names: tuple[str, ...]) -> bool:
    if not BookingRequestStatus:
        return False
    if not _model_has_field(PatientBookingRequest, "status"):
        return False

    for name in target_names:
        if hasattr(BookingRequestStatus, name):
            try:
                br.status = getattr(BookingRequestStatus, name)
                br.save(update_fields=["status"])
                return True
            except Exception:
                return False
    return False


def _booking_request_is_processed(br: PatientBookingRequest) -> bool:
    if not (BookingRequestStatus and _model_has_field(PatientBookingRequest, "status")):
        return False

    cur = getattr(br, "status", None)
    if cur is None:
        return False

    processed = []
    for nm in ("CONFIRMED", "APPROVED", "ACCEPTED", "REJECTED", "DECLINED", "CANCELLED"):
        if hasattr(BookingRequestStatus, nm):
            processed.append(getattr(BookingRequestStatus, nm))

    return cur in processed


def _parse_int(value: str | None) -> int | None:
    if not value:
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _ensure_patient_from_booking_request(br: PatientBookingRequest) -> Patient | None:
    """
    If no linked patient is found, create one from booking request data.
    """
    try:
        doc = getattr(br, "doctor", None)
        if doc is None:
            return None

        full_name = (getattr(br, "full_name", None) or "Patient").strip()
        contact = (getattr(br, "contact_info", None) or "").strip()
        dob = getattr(br, "date_of_birth", None)

        p = Patient()

        if _model_has_field(Patient, "doctor"):
            p.doctor = doc

        if _model_has_field(Patient, "full_name"):
            p.full_name = full_name

        if contact:
            if "@" in contact and _model_has_field(Patient, "email"):
                p.email = contact
            else:
                if _model_has_field(Patient, "mobile"):
                    p.mobile = contact
                elif _model_has_field(Patient, "phone"):
                    p.phone = contact

        if dob and _model_has_field(Patient, "date_of_birth"):
            p.date_of_birth = dob

        if _model_has_field(PatientBookingRequest, "user"):
            u = getattr(br, "user", None)
            if u is not None and _model_has_field(Patient, "user"):
                p.user = u  # type: ignore[attr-defined]

        p.save()
        return p
    except Exception:
        return None


# ------------------------------------------------------------------#
#               Public booking (NO LOGIN)                            #
# ------------------------------------------------------------------#
class _PublicBookingForm(forms.Form):
    doctor = forms.ModelChoiceField(queryset=Doctor.objects.none(), required=True)
    full_name = forms.CharField(max_length=200, required=True)
    contact_info = forms.CharField(max_length=200, required=True)
    date_of_birth = forms.DateField(required=False, widget=forms.DateInput(attrs={"type": "date"}))
    scheduled_time = forms.DateTimeField(
        widget=DateTimeLocalInput(),
        input_formats=DATETIME_INPUT_FORMATS,
        required=True,
    )

    def __init__(self, *args, locked_doctor: Doctor | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self.locked_doctor = locked_doctor

        base_qs = Doctor.objects.select_related("user").all()
        if locked_doctor is not None:
            base_qs = base_qs.filter(pk=locked_doctor.pk)
        self.fields["doctor"].queryset = base_qs

        if locked_doctor is not None:
            self.fields["doctor"].initial = locked_doctor.pk
            try:
                self.fields["doctor"].disabled = True
            except Exception:
                pass
            self.fields["doctor"].required = False
            self.fields["doctor"].widget.attrs.update({"disabled": "disabled"})
            try:
                self.fields["doctor"].empty_label = None
            except Exception:
                pass

    def clean(self):
        cleaned = super().clean()

        st = cleaned.get("scheduled_time")
        if st:
            st_norm = _normalize_dt(st)
            if _is_past(st_norm):
                self.add_error("scheduled_time", "يرجى اختيار وقت مستقبلي.")
            cleaned["scheduled_time"] = st_norm

        if self.locked_doctor is not None:
            cleaned["doctor"] = self.locked_doctor

        return cleaned


@require_http_methods(["GET", "POST"])
def book_appointment_public(request: HttpRequest, doctor_id: int | None = None):
    """
    Public self-booking:
    - Creates PatientBookingRequest (secretary approves).
    """
    if doctor_id is None:
        doctor_id = _parse_int(request.GET.get("doctor_id") or request.GET.get("doctor"))

    locked_doctor = None
    if doctor_id is not None:
        locked_doctor = get_object_or_404(Doctor.objects.select_related("user"), pk=doctor_id)

    if request.method == "POST":
        form = _PublicBookingForm(request.POST, locked_doctor=locked_doctor)
        if form.is_valid():
            doc: Doctor = form.cleaned_data["doctor"]
            full_name = (form.cleaned_data["full_name"] or "").strip()
            contact = (form.cleaned_data["contact_info"] or "").strip()
            dob = form.cleaned_data.get("date_of_birth")
            sched = _normalize_dt(form.cleaned_data["scheduled_time"])

            # Optional conflict block
            block_conflicts = bool(getattr(settings, "BOOKING_REQUEST_BLOCK_CONFLICTS", True))
            if block_conflicts:
                try:
                    conflict_qs = _active_appts_qs().filter(doctor=doc, scheduled_time=sched)
                    if hasattr(AppointmentStatus, "CANCELLED"):
                        conflict_qs = conflict_qs.exclude(status=AppointmentStatus.CANCELLED)
                    if conflict_qs.exists():
                        messages.error(request, "❌ هذا التوقيت محجوز مسبقًا لهذا الطبيب. يرجى اختيار وقت آخر.")
                        return render(request, "appointments/book_appointment.html", {"form": form, "doctor": doc})
                except Exception:
                    pass

            br_kwargs: dict[str, object] = {"doctor": doc}
            if _model_has_field(PatientBookingRequest, "full_name"):
                br_kwargs["full_name"] = full_name
            if _model_has_field(PatientBookingRequest, "contact_info"):
                br_kwargs["contact_info"] = contact
            if dob and _model_has_field(PatientBookingRequest, "date_of_birth"):
                br_kwargs["date_of_birth"] = dob
            if _model_has_field(PatientBookingRequest, "scheduled_time"):
                br_kwargs["scheduled_time"] = sched

            if BookingRequestStatus and _model_has_field(PatientBookingRequest, "status"):
                status_val = None
                for nm in ("REQUESTED", "PENDING"):
                    if hasattr(BookingRequestStatus, nm):
                        status_val = getattr(BookingRequestStatus, nm)
                        break
                if status_val is not None:
                    br_kwargs["status"] = status_val

            br = PatientBookingRequest.objects.create(**br_kwargs)

            _audit(
                request=request,
                action="create",
                instance=br,
                message="Public booking request created",
                extra_data={
                    "doctor_id": getattr(doc, "id", None),
                    "doctor_name": _doctor_name(doc),
                    "full_name": full_name,
                    "scheduled_time": sched.isoformat() if sched else None,
                    "source": "public_booking",
                },
            )

            try:
                if _notif_has_related_request():
                    if not Notification.objects.filter(related_booking_request=br).exists():
                        Notification.objects.create(
                            title="New booking request",
                            message=f"{full_name} requested {_doctor_name(doc)} at {_fmt_dt(sched, '%Y-%m-%d %H:%M')}",
                            related_booking_request=br,
                        )
            except Exception:
                pass

            messages.success(request, "✅ تم إرسال طلب الحجز بنجاح. سيتم تأكيده من قبل السكرتيرة.")
            return _redirect_with_query("appointments:book_appointment_success", query={"id": br.pk})

        messages.error(request, "⚠️ يرجى تصحيح الأخطاء في الحقول وإعادة المحاولة.")
    else:
        form = _PublicBookingForm(locked_doctor=locked_doctor)

    ctx = {"form": form, "doctor": locked_doctor}
    try:
        return render(request, "appointments/book_appointment.html", ctx)
    except TemplateDoesNotExist:
        return HttpResponse("Public booking page is available.", status=200)


@require_GET
def book_appointment_success(request: HttpRequest):
    booking_id = (request.GET.get("id") or "").strip()
    ctx = {"booking_id": booking_id}
    try:
        return render(request, "appointments/book_success.html", ctx)
    except TemplateDoesNotExist:
        return HttpResponse("Booking request submitted successfully.", status=200)


# ------------------------------------------------------------------#
#                     Secretary Dashboard                            #
# ------------------------------------------------------------------#
@secretary_required
@require_GET
def secretary_dashboard(request: HttpRequest):
    today = _today()

    base = _active_appts_qs().select_related("patient", "doctor__user")
    base = _filter_appointments_for_user(base, request.user)

    todays_all = _filter_by_day(base, today).order_by("scheduled_time")
    todays_queue = todays_all.filter(status__in=_queue_active_statuses())

    revenue_today = 0
    if _model_has_field(Appointment, "iqd_amount"):
        revenue_today = todays_all.aggregate(total=Sum("iqd_amount")).get("total") or 0

    assigned_doctor = _secretary_assigned_doctor(request.user)
    pat_qs = Patient.objects.all()
    if assigned_doctor is not None and _model_has_field(Patient, "doctor"):
        pat_qs = pat_qs.filter(doctor=assigned_doctor)

    stats = {
        "patients_today": todays_all.values("patient_id").distinct().count(),
        "new_patients_today": (
            pat_qs.filter(created_at__date=today).count() if _model_has_field(Patient, "created_at") else 0
        ),
        "total_patients": pat_qs.count(),
        "appointments_today": todays_all.count(),
        "revenue_today_iqd": revenue_today,
    }

    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    week_qs = _filter_by_day_range(base, week_start, week_end)

    if _model_has_field(Appointment, "scheduled_day"):
        rows = week_qs.values("scheduled_day").annotate(count=Count("id"))
        counts = {r["scheduled_day"]: r["count"] for r in rows}
    else:
        rows = week_qs.annotate(day=TruncDate("scheduled_time")).values("day").annotate(count=Count("id"))
        counts = {r["day"]: r["count"] for r in rows}

    labels = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    chart = [counts.get(week_start + timedelta(days=i), 0) for i in range(7)]

    call_next_url = ""
    assigned_doctor_id = None
    if assigned_doctor is not None:
        assigned_doctor_id = assigned_doctor.id
        call_next_url = reverse("appointments:call_next_api", kwargs={"doctor_id": assigned_doctor.id})

    ctx = {
        "appointment_form": AppointmentForm(),
        "patient_form": SecretaryPatientForm(),
        "appointments": base.order_by("-scheduled_time")[:20],
        "today_appointments": todays_queue,
        "today_appointments_all": todays_all,
        "stats": stats,
        "chart_data_json": json.dumps({"labels": labels, "data": chart}),
        "assigned_doctor": assigned_doctor,
        "assigned_doctor_id": assigned_doctor_id,
        "call_next_url": call_next_url,
        "queue_api_url": reverse("appointments:queue_number_api"),
        "recycle_bin_url": reverse("appointments:appointment_recycle_bin"),
    }
    return render(request, "appointments/secretary_dashboard.html", ctx)


# ------------------------------------------------------------------#
#                        Appointment CRUD                            #
# ------------------------------------------------------------------#
@secretary_required
@require_http_methods(["GET", "POST"])
def create_appointment(request: HttpRequest):
    assigned_doctor = _secretary_assigned_doctor(request.user)
    form = AppointmentForm(request.POST or None)

    _lock_form_doctor_field(form, assigned_doctor)

    if request.method == "GET" and "status" in form.fields:
        try:
            form.fields["status"].initial = _secretary_default_status()
        except Exception:
            pass

    if request.method == "POST" and form.is_valid():
        appt: Appointment = form.save(commit=False)

        if assigned_doctor is not None:
            appt.doctor = assigned_doctor

        appt.scheduled_time = _normalize_dt(getattr(appt, "scheduled_time", None))

        try:
            appt.save()
            if hasattr(form, "save_m2m"):
                form.save_m2m()
        except IntegrityError:
            messages.error(request, "❌ هذا التوقيت محجوز مسبقًا لهذا الطبيب. يرجى اختيار وقت آخر.")
            return render(request, "appointments/create_appointment.html", {"form": form})
        except ValidationError as e:
            messages.error(
                request,
                _extract_validation_message(
                    e,
                    default="❌ لا يمكن حفظ الموعد. يرجى التحقق من البيانات.",
                    preferred_field="scheduled_time",
                ),
            )
            return render(request, "appointments/create_appointment.html", {"form": form})

        _audit(
            request=request,
            action="create",
            instance=appt,
            message="Appointment created",
            extra_data={
                "appointment_id": appt.pk,
                "patient_id": getattr(appt, "patient_id", None),
                "patient_name": getattr(getattr(appt, "patient", None), "full_name", ""),
                "doctor_id": getattr(appt, "doctor_id", None),
                "doctor_name": _doctor_name(getattr(appt, "doctor", None)),
                "scheduled_time": appt.scheduled_time.isoformat() if getattr(appt, "scheduled_time", None) else None,
                "status": str(getattr(appt, "status", "")),
                "queue_number": getattr(appt, "queue_number", None),
            },
        )

        time_part = _fmt_dt(appt.scheduled_time, "%I:%M %p") if appt.scheduled_time else "—"
        qno = (
            f" | رقم الدور: {_format_queue_number(getattr(appt, 'queue_number', None))}"
            if getattr(appt, "queue_number", None)
            else ""
        )

        messages.success(
            request,
            (
                f"✅ تم حجز الموعد بنجاح للمريض: {appt.patient.full_name} "
                f"مع الدكتور/ة: {_doctor_name(appt.doctor)} "
                f"في الساعة {time_part}.{qno}"
            ),
        )
        return _redirect_with_query("appointments:appointment_list", query={"created": appt.pk})

    if request.method == "POST":
        messages.error(request, "⚠️ لم يتم حفظ الموعد. يرجى تصحيح الأخطاء في الحقول وإعادة المحاولة.")

    return render(request, "appointments/create_appointment.html", {"form": form})


@staff_ticket_required
@require_GET
def appointment_ticket(request: HttpRequest, pk: int):
    qs = _active_appts_qs().select_related("doctor__user", "patient")
    qs = _filter_appointments_for_user(qs, request.user)
    appt = get_object_or_404(qs, pk=pk)

    try:
        import qrcode  # type: ignore
    except Exception:
        messages.error(request, "⚠️ مكتبة QR غير مثبتة (qrcode). ثبتيها: pip install qrcode[pil]")
        return redirect("appointments:appointment_list")

    qr = qrcode.make(request.build_absolute_uri(), box_size=6, border=2)
    buf = io.BytesIO()
    qr.save(buf, format="PNG")

    _audit(
        request=request,
        action="view",
        instance=appt,
        message="Appointment ticket viewed",
        extra_data={"appointment_id": appt.pk},
    )

    ctx = {
        "appointment": appt,
        "doctor_name": _doctor_name(appt.doctor),
        "doctor_spec": _doctor_specialty(appt.doctor),
        "doctor_room": _doctor_room(appt.doctor),
        "secretary_name": _user_name(request.user),
        "qr_code": base64.b64encode(buf.getvalue()).decode(),
    }
    return render(request, "appointments/appointment_ticket.html", ctx)


@secretary_required
@require_http_methods(["GET", "POST"])
def edit_appointment(request: HttpRequest, pk: int):
    qs = _filter_appointments_for_user(_active_appts_qs(), request.user)
    appt = get_object_or_404(qs, pk=pk)

    form = AppointmentForm(request.POST or None, instance=appt)

    assigned_doctor = _secretary_assigned_doctor(request.user)
    _lock_form_doctor_field(form, assigned_doctor)

    if request.method == "POST" and form.is_valid():
        changed_fields = list(getattr(form, "changed_data", []) or [])

        old_snapshot = {
            "scheduled_time": appt.scheduled_time.isoformat() if getattr(appt, "scheduled_time", None) else None,
            "status": str(getattr(appt, "status", "")),
            "doctor_id": getattr(appt, "doctor_id", None),
            "patient_id": getattr(appt, "patient_id", None),
            "queue_number": getattr(appt, "queue_number", None),
        }

        appt = form.save(commit=False)

        if assigned_doctor is not None:
            appt.doctor = assigned_doctor

        appt.scheduled_time = _normalize_dt(getattr(appt, "scheduled_time", None))

        try:
            appt.save()
            if hasattr(form, "save_m2m"):
                form.save_m2m()
        except IntegrityError:
            messages.error(request, "❌ هذا التوقيت محجوز مسبقًا لهذا الطبيب. يرجى اختيار وقت آخر.")
            return render(request, "appointments/edit_appointment.html", {"form": form, "appointment": appt})
        except ValidationError as e:
            messages.error(
                request,
                _extract_validation_message(
                    e,
                    default="❌ لا يمكن تحديث الموعد. يرجى التحقق من البيانات.",
                    preferred_field="scheduled_time",
                ),
            )
            return render(request, "appointments/edit_appointment.html", {"form": form, "appointment": appt})

        _audit(
            request=request,
            action="update",
            instance=appt,
            message="Appointment updated",
            extra_data={
                "appointment_id": appt.pk,
                "changed_fields": changed_fields,
                "before": old_snapshot,
                "after": {
                    "scheduled_time": appt.scheduled_time.isoformat() if getattr(appt, "scheduled_time", None) else None,
                    "status": str(getattr(appt, "status", "")),
                    "doctor_id": getattr(appt, "doctor_id", None),
                    "patient_id": getattr(appt, "patient_id", None),
                    "queue_number": getattr(appt, "queue_number", None),
                },
            },
        )

        messages.success(request, "✅ تم تحديث بيانات الموعد بنجاح.")
        return redirect("appointments:appointment_list")

    if request.method == "POST":
        messages.error(request, "⚠️ لم يتم تحديث الموعد. يرجى مراجعة البيانات المدخلة.")

    return render(request, "appointments/edit_appointment.html", {"form": form, "appointment": appt})


@secretary_required
@require_http_methods(["GET", "POST"])
def cancel_appointment(request: HttpRequest, pk: int):
    """
    Cancel = تغيير الحالة إلى CANCELLED (ليس حذف).
    """
    qs = _filter_appointments_for_user(_active_appts_qs(), request.user)
    appt = get_object_or_404(qs, pk=pk)

    if getattr(appt, "status", None) == getattr(AppointmentStatus, "COMPLETED", None):
        messages.error(request, "❌ لا يمكن إلغاء موعد تم الانتهاء منه بالفعل.")
        return redirect("appointments:appointment_list")

    if request.method == "POST":
        reason = (request.POST.get("reason") or "").strip()

        update_kwargs: dict[str, object] = {"status": AppointmentStatus.CANCELLED}

        if _model_has_field(Appointment, "queue_number"):
            update_kwargs["queue_number"] = None

        if reason and _model_has_field(Appointment, "notes"):
            stamp_dt = _now_local()
            stamp = stamp_dt.strftime("%Y-%m-%d %H:%M") if stamp_dt else timezone.now().strftime("%Y-%m-%d %H:%M")
            user_display = request.user.get_full_name() or request.user.username
            note_line = f"[Cancelled {stamp} by {user_display}] {reason}"
            existing_notes = getattr(appt, "notes", "") or ""
            new_notes = f"{existing_notes}\n{note_line}".strip()
            update_kwargs["notes"] = new_notes

        Appointment.objects.filter(pk=appt.pk).update(**update_kwargs)

        try:
            appt.status = AppointmentStatus.CANCELLED
            if "queue_number" in update_kwargs:
                appt.queue_number = None  # type: ignore[attr-defined]
            if "notes" in update_kwargs:
                appt.notes = update_kwargs["notes"]  # type: ignore[attr-defined]
        except Exception:
            pass

        _audit(
            request=request,
            action="update",
            instance=appt,
            message="Appointment cancelled",
            extra_data={
                "appointment_id": appt.pk,
                "reason": reason,
                "new_status": str(AppointmentStatus.CANCELLED),
            },
        )

        messages.success(request, "✅ تم إلغاء الموعد بنجاح.")
        return redirect("appointments:appointment_list")

    return render(request, "appointments/cancel_confirmation.html", {"appointment": appt})


@secretary_required
@require_http_methods(["GET", "POST"])
def delete_appointment(request: HttpRequest, pk: int):
    """
    Soft Delete (Recycle Bin)
    """
    qs = _filter_appointments_for_user(_active_appts_qs(), request.user)
    appt = get_object_or_404(qs, pk=pk)

    if request.method == "POST":
        _audit(
            request=request,
            action="delete",
            instance=appt,
            message="Appointment moved to recycle bin",
            extra_data={
                "appointment_id": appt.pk,
                "soft_delete": True,
                "patient_id": getattr(appt, "patient_id", None),
                "doctor_id": getattr(appt, "doctor_id", None),
                "scheduled_time": appt.scheduled_time.isoformat() if getattr(appt, "scheduled_time", None) else None,
            },
        )

        try:
            appt.delete(user=request.user)  # type: ignore[arg-type]
        except TypeError:
            appt.delete()

        messages.success(request, "🗑️ تم نقل الموعد إلى سلة المحذوفات (Recycle Bin).")
        return redirect("appointments:appointment_list")

    return render(request, "appointments/delete_confirmation.html", {"appointment": appt})


# ------------------------------------------------------------------#
#                 Recycle Bin (Appointments)                         #
# ------------------------------------------------------------------#
@secretary_required
@require_GET
def appointment_recycle_bin(request: HttpRequest):
    if not hasattr(Appointment, "deleted_objects"):
        messages.error(request, "⚠️ سلة المحذوفات غير مفعلة في هذا الإصدار.")
        return redirect("appointments:appointment_list")

    qs = Appointment.deleted_objects.select_related("patient", "doctor__user")  # type: ignore[attr-defined]

    order_fields: list[str] = []
    if _model_has_field(Appointment, "deleted_at"):
        order_fields.append("-deleted_at")
    elif _model_has_field(Appointment, "updated_at"):
        order_fields.append("-updated_at")
    order_fields.append("-pk")
    qs = qs.order_by(*order_fields)

    qs = _filter_appointments_for_user(qs, request.user)

    q = (request.GET.get("q") or "").strip()
    if q:
        notes_q = Q()
        if _model_has_field(Appointment, "notes"):
            notes_q = Q(notes__icontains=q)
        qs = qs.filter(
            Q(patient__full_name__icontains=q)
            | Q(doctor__user__first_name__icontains=q)
            | Q(doctor__user__last_name__icontains=q)
            | notes_q
        )

    page = Paginator(qs, 20).get_page(request.GET.get("page"))
    return render(
        request,
        "appointments/appointment_recycle_bin.html",
        {"deleted_appointments": page, "search_query": q},
    )


@secretary_required
@require_POST
def restore_appointment(request: HttpRequest, pk: int):
    if not hasattr(Appointment, "all_objects"):
        messages.error(request, "⚠️ الاسترجاع غير متاح (all_objects غير موجود).")
        return redirect("appointments:appointment_recycle_bin")

    qs = Appointment.all_objects.select_related("patient", "doctor__user")  # type: ignore[attr-defined]
    qs = _filter_appointments_for_user(qs, request.user)
    appt = get_object_or_404(qs, pk=pk)

    is_deleted = False
    if hasattr(appt, "is_deleted"):
        try:
            is_deleted = bool(getattr(appt, "is_deleted"))
        except Exception:
            is_deleted = False
    if not is_deleted and hasattr(appt, "deleted_at"):
        try:
            is_deleted = bool(getattr(appt, "deleted_at"))
        except Exception:
            is_deleted = False

    if not is_deleted:
        messages.info(request, "ℹ️ هذا الموعد ليس ضمن سلة المحذوفات.")
        return redirect("appointments:appointment_recycle_bin")

    try:
        if hasattr(appt, "restore"):
            try:
                appt.restore(user=request.user)  # type: ignore[misc]
            except TypeError:
                appt.restore()
        else:
            if hasattr(appt, "is_deleted"):
                setattr(appt, "is_deleted", False)
                appt.save(update_fields=["is_deleted"])
            elif hasattr(appt, "deleted_at"):
                setattr(appt, "deleted_at", None)
                appt.save(update_fields=["deleted_at"])

        _audit(
            request=request,
            action="update",
            instance=appt,
            message="Appointment restored from recycle bin",
            extra_data={"appointment_id": appt.pk, "restored": True},
        )

        messages.success(request, "✅ تم استرجاع الموعد بنجاح.")
    except IntegrityError:
        messages.error(
            request,
            "❌ لا يمكن استرجاع الموعد لأن هناك تعارض (نفس الطبيب ونفس الوقت موجود). "
            "رجاءً غيّري وقت الموعد أو احذفي الموعد المتعارض.",
        )
    except ValidationError as e:
        msg = "❌ لا يمكن استرجاع الموعد. يرجى التحقق من البيانات."
        if getattr(e, "messages", None):
            msg = f"❌ {e.messages[0]}"
        messages.error(request, msg)

    return redirect("appointments:appointment_recycle_bin")


@secretary_required
@require_http_methods(["GET", "POST"])
def hard_delete_appointment(request: HttpRequest, pk: int):
    """Permanent delete (SUPERUSER ONLY) from Recycle Bin."""
    if not request.user.is_superuser:
        raise PermissionDenied("Hard delete is restricted to administrators only.")

    qs = (
        Appointment.all_objects.select_related("patient", "doctor__user")  # type: ignore[attr-defined]
        if hasattr(Appointment, "all_objects")
        else _active_appts_qs().select_related("patient", "doctor__user")
    )
    appt = get_object_or_404(qs, pk=pk)

    if request.method == "POST":
        _audit(
            request=request,
            action="delete",
            instance=appt,
            message="Appointment permanently deleted",
            extra_data={
                "appointment_id": appt.pk,
                "hard_delete": True,
                "patient_id": getattr(appt, "patient_id", None),
                "doctor_id": getattr(appt, "doctor_id", None),
                "scheduled_time": appt.scheduled_time.isoformat() if getattr(appt, "scheduled_time", None) else None,
            },
        )

        try:
            appt.delete(hard=True)  # type: ignore[arg-type]
        except TypeError:
            appt.delete()

        messages.success(request, "🗑️ تم حذف الموعد نهائيًا (Permanent Delete).")
        return redirect("appointments:appointment_recycle_bin")

    return render(request, "appointments/delete_confirmation.html", {"appointment": appt})


@staff_ticket_required
@require_GET
def appointment_list(request: HttpRequest):
    sort = request.GET.get("sort", "scheduled_time")
    fld = {
        "patient": "patient__full_name",
        "doctor": "doctor__user__first_name",
        "scheduled_time": "scheduled_time",
    }.get(sort, "scheduled_time")

    qs = _active_appts_qs().select_related("patient", "doctor__user")
    qs = _filter_appointments_for_user(qs, request.user)

    status_key = (request.GET.get("status") or "all").lower()
    status_map = {
        "pending": AppointmentStatus.PENDING,
        "completed": AppointmentStatus.COMPLETED,
        "cancelled": AppointmentStatus.CANCELLED,
    }
    if hasattr(AppointmentStatus, "CALLED"):
        status_map["called"] = getattr(AppointmentStatus, "CALLED")
    for name, key in (("APPROVED", "approved"), ("CONFIRMED", "confirmed"), ("ACCEPTED", "accepted")):
        if hasattr(AppointmentStatus, name):
            status_map[key] = getattr(AppointmentStatus, name)

    if status_key in status_map:
        qs = qs.filter(status=status_map[status_key])

    q = (request.GET.get("q") or "").strip()
    if q:
        notes_q = Q()
        if _model_has_field(Appointment, "notes"):
            notes_q = Q(notes__icontains=q)
        qs = qs.filter(
            Q(patient__full_name__icontains=q)
            | Q(doctor__user__first_name__icontains=q)
            | Q(doctor__user__last_name__icontains=q)
            | notes_q
        )

    created_id = request.GET.get("created")
    created_int: int | None = None
    if created_id:
        try:
            created_int = int(created_id)
        except Exception:
            created_int = None

    direction = (request.GET.get("dir") or "desc").lower()
    if direction not in {"asc", "desc"}:
        direction = "desc"
    order_field = fld if direction == "asc" else f"-{fld}"

    page = Paginator(qs.order_by(order_field), 10).get_page(request.GET.get("page"))
    return render(
        request,
        "appointments/appointment_list.html",
        {
            "appointments": page,
            "search_query": q,
            "current_sort": sort,
            "current_dir": direction,
            "current_status": status_key,
            "created_id": created_int,
            "recycle_bin_url": reverse("appointments:appointment_recycle_bin"),
        },
    )


# ------------------------------------------------------------------#
#                Patient Portal Booking (IN-APP)                     #
# ------------------------------------------------------------------#
class _PatientPortalBookingForm(forms.ModelForm):
    scheduled_time = forms.DateTimeField(
        widget=DateTimeLocalInput(),
        input_formats=DATETIME_INPUT_FORMATS,
        required=True,
    )

    class Meta:
        model = Appointment
        fields = ["scheduled_time"]

    def __init__(self, *args, doctor: Doctor | None = None, **kwargs):
        self.doctor = doctor
        super().__init__(*args, **kwargs)

    def clean(self):
        cleaned = super().clean()
        st = cleaned.get("scheduled_time")
        if st:
            st_norm = _normalize_dt(st)
            if _is_past(st_norm):
                self.add_error("scheduled_time", "Please choose a future time.")
            cleaned["scheduled_time"] = st_norm
        return cleaned


@login_required
@require_http_methods(["GET", "POST"])
def book_patient(request: HttpRequest, doctor_id: int):
    if not is_patient(request.user):
        return HttpResponseForbidden("Patients only.")

    doctor = get_object_or_404(Doctor, pk=doctor_id)
    patient = _get_patient_for_user(request.user)
    if not patient:
        return HttpResponseForbidden("Patient profile not found.")

    if request.method == "POST":
        form = _PatientPortalBookingForm(request.POST, doctor=doctor)
        if form.is_valid():
            sched = _normalize_dt(form.cleaned_data["scheduled_time"])

            # Prefer booking request workflow if BookingRequestStatus exists
            if BookingRequestStatus:
                br_kwargs: dict[str, object] = {"doctor": doctor}

                full_name = patient.full_name or request.user.get_full_name() or request.user.username
                contact = getattr(patient, "phone", "") or getattr(patient, "mobile", "") or request.user.email or ""
                dob = getattr(patient, "date_of_birth", None)

                if _model_has_field(PatientBookingRequest, "full_name"):
                    br_kwargs["full_name"] = full_name
                if _model_has_field(PatientBookingRequest, "contact_info"):
                    br_kwargs["contact_info"] = contact
                if _model_has_field(PatientBookingRequest, "date_of_birth"):
                    br_kwargs["date_of_birth"] = dob
                if _model_has_field(PatientBookingRequest, "scheduled_time"):
                    br_kwargs["scheduled_time"] = sched

                status_val = None
                for nm in ("REQUESTED", "PENDING"):
                    if hasattr(BookingRequestStatus, nm):
                        status_val = getattr(BookingRequestStatus, nm)
                        break
                if status_val is not None and _model_has_field(PatientBookingRequest, "status"):
                    br_kwargs["status"] = status_val

                if _model_has_field(PatientBookingRequest, "patient"):
                    br_kwargs["patient"] = patient
                if _model_has_field(PatientBookingRequest, "user"):
                    br_kwargs["user"] = request.user

                br = PatientBookingRequest.objects.create(**br_kwargs)

                _audit(
                    request=request,
                    action="create",
                    instance=br,
                    message="Patient portal booking request created",
                    extra_data={
                        "doctor_id": getattr(doctor, "id", None),
                        "doctor_name": _doctor_name(doctor),
                        "patient_id": getattr(patient, "id", None),
                        "patient_name": getattr(patient, "full_name", ""),
                        "scheduled_time": sched.isoformat() if sched else None,
                        "source": "patient_portal",
                    },
                )

                try:
                    if _notif_has_related_request():
                        if not Notification.objects.filter(related_booking_request=br).exists():
                            Notification.objects.create(
                                title="New booking request",
                                message=f"{full_name} requested {_doctor_name(doctor)} at {_fmt_dt(sched, '%Y-%m-%d %H:%M')}",
                                related_booking_request=br,
                            )
                except Exception:
                    pass

                messages.success(request, "✅ تم إرسال طلبك وسيتم تأكيده من قبل السكرتيرة.")
                return redirect(_safe_reverse("patient:dashboard", default=reverse("appointments:my_appointments")))

            # Direct appointment fallback
            appt = Appointment(
                patient=patient,
                doctor=doctor,
                scheduled_time=sched,
                status=AppointmentStatus.PENDING,
            )
            try:
                appt.save()
            except IntegrityError:
                messages.error(request, "❌ This time slot is already booked for this doctor.")
                return redirect("appointments:my_appointments")
            except ValidationError as e:
                msg = "❌ Cannot create appointment."
                if getattr(e, "messages", None):
                    msg = f"❌ {e.messages[0]}"
                messages.error(request, msg)
                return redirect("appointments:my_appointments")

            _audit(
                request=request,
                action="create",
                instance=appt,
                message="Appointment created from patient portal",
                extra_data={
                    "appointment_id": appt.pk,
                    "patient_id": getattr(patient, "id", None),
                    "doctor_id": getattr(doctor, "id", None),
                    "scheduled_time": sched.isoformat() if sched else None,
                    "source": "patient_portal_direct",
                },
            )

            try:
                Notification.objects.create(
                    title="New appointment",
                    message=f"{patient.full_name} booked {_doctor_name(doctor)} at {_fmt_dt(sched, '%Y-%m-%d %H:%M')}",
                )
            except Exception:
                pass

            messages.success(request, "✅ تم إرسال طلبك وهو بانتظار التأكيد.")
            return redirect("appointments:my_appointments")

        messages.error(request, "❌ Please correct the errors in the selected date/time.")
    else:
        form = _PatientPortalBookingForm(doctor=doctor)

    return render(request, "appointments/book_patient.html", {"form": form, "doctor": doctor, "patient": patient})


@login_required
@require_GET
def my_appointments(request: HttpRequest):
    if not is_patient(request.user):
        return HttpResponseForbidden("Patients only.")

    patient = _get_patient_for_user(request.user)
    if not patient:
        return HttpResponseForbidden("Patient profile not found.")

    appointments = (
        _active_appts_qs()
        .filter(patient=patient)
        .select_related("doctor", "doctor__user")
        .order_by("-scheduled_time")
    )

    booking_requests: list[PatientBookingRequest] = []
    if BookingRequestStatus:
        q = PatientBookingRequest.objects.all()

        if _model_has_field(PatientBookingRequest, "patient"):
            q = q.filter(patient=patient)
        elif _model_has_field(PatientBookingRequest, "user"):
            q = q.filter(user=request.user)
        else:
            lookups = Q()
            phone = getattr(patient, "phone", None) or getattr(patient, "mobile", None)
            if phone and _model_has_field(PatientBookingRequest, "contact_info"):
                lookups |= Q(contact_info__icontains=str(phone))
            if request.user.email and _model_has_field(PatientBookingRequest, "contact_info"):
                lookups |= Q(contact_info__icontains=request.user.email)
            display_name = patient.full_name or request.user.get_full_name() or request.user.username
            if display_name and _model_has_field(PatientBookingRequest, "full_name"):
                lookups |= Q(full_name__icontains=display_name)
            if lookups:
                q = q.filter(lookups)

        has_submitted_at = _model_has_field(PatientBookingRequest, "submitted_at")
        q = q.select_related("doctor", "doctor__user").order_by(
            "-submitted_at" if has_submitted_at else "-scheduled_time"
        )
        booking_requests = list(q[:50])

    return render(
        request,
        "appointments/my_appointments.html",
        {"appointments": appointments, "booking_requests": booking_requests},
    )


# ------------------------------------------------------------------#
#           Approving appointments / booking requests                #
# ------------------------------------------------------------------#
@secretary_required
@require_POST
def approve_appointment(request: HttpRequest, pk: int):
    return confirm_appointment(request, pk)


@secretary_required
@require_POST
def confirm_appointment(request: HttpRequest, pk: int):
    qs = _filter_appointments_for_user(_active_appts_qs(), request.user)
    appt = get_object_or_404(qs, pk=pk)

    old_status = str(getattr(appt, "status", ""))

    if getattr(appt, "status", None) == getattr(AppointmentStatus, "CANCELLED", object()):
        messages.error(request, "❌ لا يمكن تأكيد موعد تم إلغاؤه.")
    elif getattr(appt, "status", None) == getattr(AppointmentStatus, "COMPLETED", object()):
        messages.error(request, "❌ لا يمكن تأكيد موعد مكتمل. يرجى إنشاء موعد جديد إذا لزم الأمر.")
    else:
        appt.status = _secretary_default_status()
        appt.save()

        _audit(
            request=request,
            action="update",
            instance=appt,
            message="Appointment confirmed/approved",
            extra_data={
                "appointment_id": appt.pk,
                "old_status": old_status,
                "new_status": str(getattr(appt, "status", "")),
            },
        )

        messages.success(request, "✅ تم تأكيد الموعد وتحديث حالته في نظام ClinicHub.")

    return redirect("appointments:appointment_list")


@secretary_required
@require_http_methods(["GET", "POST"])
def approve_booking_request(request: HttpRequest, pk: int):
    """
    Approve MUST create ONE appointment at the requested slot (doctor + scheduled_time)
    and be idempotent + concurrency-safe.

    Fix:
    - Lock only the booking request row (avoid nullable joins with FOR UPDATE).
    - Link booking request -> patient.
    - Force Patient.doctor_id = BookingRequest.doctor_id if Patient has doctor field.
    """
    fallback_next = reverse("appointments:booking_requests_list")
    next_url = _safe_next_url(request, fallback=fallback_next)

    if request.method != "POST":
        return redirect(next_url)

    cancelled_status = getattr(AppointmentStatus, "CANCELLED", None)

    try:
        with transaction.atomic():
            # Lock only booking request row
            br_qs = PatientBookingRequest.objects.all()
            try:
                br_qs = br_qs.select_for_update(of=("self",))
            except TypeError:
                br_qs = br_qs.select_for_update()

            br_qs = br_qs.select_related("doctor", "doctor__user")
            br_qs = _filter_booking_requests_for_user(br_qs, request.user)
            br = get_object_or_404(br_qs, pk=pk)

            if getattr(br, "doctor", None) is None:
                messages.error(request, "⚠️ لا يمكن اعتماد الطلب لأن الطبيب غير محدد في الطلب.")
                return redirect(next_url)

            if _booking_request_is_processed(br):
                _mark_related_notifications_read(br)
                _mark_booking_request_seen(br)
                messages.info(request, "ℹ️ هذا الطلب تمّت معالجته مسبقًا.")
                return redirect(next_url)

            # Resolve / create patient
            patient_obj: Patient | None = None

            if _model_has_field(PatientBookingRequest, "patient") and getattr(br, "patient", None):
                patient_obj = br.patient  # type: ignore[attr-defined]
            elif (
                _model_has_field(PatientBookingRequest, "user")
                and getattr(br, "user", None)
                and _model_has_field(Patient, "user")
            ):
                patient_obj = Patient.objects.filter(user=br.user).first()  # type: ignore[attr-defined]

            if not patient_obj:
                qs_pat = Patient.objects.all()
                contact = getattr(br, "contact_info", None)

                if contact:
                    look = Q()
                    if _model_has_field(Patient, "mobile"):
                        look |= Q(mobile__icontains=str(contact))
                    if _model_has_field(Patient, "phone"):
                        look |= Q(phone__icontains=str(contact))
                    if _model_has_field(Patient, "user") and _model_has_field(User, "email"):
                        look |= Q(user__email__iexact=str(contact))
                    if look:
                        qs_pat = qs_pat.filter(look)

                if (not qs_pat.exists()) and getattr(br, "full_name", None) and _model_has_field(Patient, "full_name"):
                    qs_pat = Patient.objects.filter(full_name__icontains=br.full_name)

                patient_obj = qs_pat.first()

            if not patient_obj:
                patient_obj = _ensure_patient_from_booking_request(br)

            if not patient_obj:
                messages.error(
                    request,
                    "⚠️ لا يمكن اعتماد الطلب: لم يتم العثور على/إنشاء سجل المريض. يرجى إنشاء/ربط المريض أولاً.",
                )
                return redirect(next_url)

            # Link booking request -> patient
            if _model_has_field(PatientBookingRequest, "patient"):
                try:
                    if getattr(br, "patient", None) is None or getattr(br, "patient_id", None) != patient_obj.id:
                        PatientBookingRequest.objects.filter(pk=br.pk).update(patient=patient_obj)
                        try:
                            br.patient = patient_obj  # type: ignore[attr-defined]
                        except Exception:
                            pass
                except Exception:
                    pass

            # Force patient doctor assignment (for scoped secretary lists)
            if _model_has_field(Patient, "doctor") and getattr(br, "doctor_id", None):
                Patient.objects.filter(pk=patient_obj.pk).update(doctor_id=br.doctor_id)
                try:
                    patient_obj.doctor_id = br.doctor_id
                except Exception:
                    pass

            scheduled_time = _normalize_dt(getattr(br, "scheduled_time", None))
            if not scheduled_time:
                messages.error(request, "⚠️ لا يمكن اعتماد الطلب لأن وقت الحجز غير موجود.")
                return redirect(next_url)

            # Idempotent: already linked
            if _model_has_field(PatientBookingRequest, "appointment"):
                existing_appt = getattr(br, "appointment", None)
                if isinstance(existing_appt, Appointment):
                    if cancelled_status is None or getattr(existing_appt, "status", None) != cancelled_status:
                        _set_booking_request_status(br, ("CONFIRMED", "APPROVED", "ACCEPTED"))
                        _mark_related_notifications_read(br)
                        _mark_booking_request_seen(br)
                        messages.success(request, "✅ تم اعتماد الطلب (الموعد كان مرتبطًا مسبقًا).")
                        return _redirect_with_query("appointments:appointment_list", query={"created": existing_appt.pk})

            # Create/reuse slot appointment (race-safe)
            appt: Appointment | None = None
            created_new_appointment = False

            slot_qs = _active_appts_qs().filter(doctor=br.doctor, scheduled_time=scheduled_time)
            try:
                slot_qs = slot_qs.select_for_update(of=("self",))
            except TypeError:
                slot_qs = slot_qs.select_for_update()

            if cancelled_status is not None:
                slot_qs = slot_qs.exclude(status=cancelled_status)

            existing_slot = slot_qs.first()

            if existing_slot:
                if existing_slot.patient_id != patient_obj.id:
                    messages.error(request, "⚠️ لا يمكن اعتماد هذا الطلب لأن هذا التوقيت محجوز لمريض آخر.")
                    return redirect(next_url)

                existing_slot.status = _secretary_default_status()
                existing_slot.save()
                appt = existing_slot
                created_new_appointment = False
            else:
                appt = Appointment(
                    patient=patient_obj,
                    doctor=br.doctor,
                    scheduled_time=scheduled_time,
                    status=_secretary_default_status(),
                )
                appt.save()
                created_new_appointment = True

            # link request -> appointment
            if _model_has_field(PatientBookingRequest, "appointment"):
                PatientBookingRequest.objects.filter(pk=br.pk).update(appointment=appt)
                try:
                    br.appointment = appt  # type: ignore[attr-defined]
                except Exception:
                    pass

            _set_booking_request_status(br, ("CONFIRMED", "APPROVED", "ACCEPTED"))
            _mark_booking_request_seen(br)
            _mark_related_notifications_read(br)

            _audit(
                request=request,
                action="create" if created_new_appointment else "update",
                instance=appt,
                message="Booking request approved and appointment linked",
                extra_data={
                    "booking_request_id": br.pk,
                    "appointment_id": appt.pk if appt else None,
                    "created_new_appointment": created_new_appointment,
                    "patient_id": getattr(appt, "patient_id", None) if appt else None,
                    "doctor_id": getattr(appt, "doctor_id", None) if appt else None,
                    "scheduled_time": appt.scheduled_time.isoformat() if appt and getattr(appt, "scheduled_time", None) else None,
                    "appointment_status": str(getattr(appt, "status", "")) if appt else "",
                },
            )

            messages.success(request, "✅ تم اعتماد طلب الحجز وإنشاء الموعد تلقائيًا بنفس الوقت المحدد.")
            return _redirect_with_query("appointments:appointment_list", query={"created": appt.pk if appt else None})

    except IntegrityError:
        messages.error(request, "⚠️ تعارض أثناء اعتماد الطلب (نفس الطبيب ونفس الوقت).")
        return redirect(next_url)
    except ValidationError as e:
        error_msg = None
        if hasattr(e, "message_dict"):
            msgs = e.message_dict.get("scheduled_time")
            if isinstance(msgs, (list, tuple)) and msgs:
                error_msg = msgs[0]
        if not error_msg:
            error_msg = "لا يمكن اعتماد هذا التوقيت لهذا الطبيب، لأنه محجوز بالفعل أو غير صالح."
        messages.error(request, f"⚠️ لم يتم اعتماد طلب الحجز: {error_msg}")
        return redirect(next_url)
    except Exception:
        if settings.DEBUG:
            raise
        messages.error(request, "❌ حدث خطأ غير متوقع أثناء اعتماد الطلب. يرجى إعادة المحاولة.")
        return redirect(next_url)


@secretary_required
@require_http_methods(["GET", "POST"])
def reject_booking_request(request: HttpRequest, pk: int):
    if request.method != "POST":
        return redirect(_safe_next_url(request, fallback=reverse("appointments:booking_requests_list")))

    br_qs = _filter_booking_requests_for_user(PatientBookingRequest.objects.all(), request.user)
    br = get_object_or_404(br_qs, pk=pk)

    next_url = _safe_next_url(request, fallback=reverse("appointments:booking_requests_list"))
    old_status = str(getattr(br, "status", "")) if _model_has_field(PatientBookingRequest, "status") else ""

    try:
        _set_booking_request_status(br, ("REJECTED", "DECLINED", "CANCELLED"))
        _mark_booking_request_seen(br)
        _mark_related_notifications_read(br)

        _audit(
            request=request,
            action="update",
            instance=br,
            message="Booking request rejected",
            extra_data={
                "booking_request_id": br.pk,
                "old_status": old_status,
                "new_status": str(getattr(br, "status", "")) if _model_has_field(PatientBookingRequest, "status") else "",
            },
        )

        messages.success(request, "✅ تم رفض طلب الحجز.")
        return redirect(next_url)
    except Exception:
        messages.error(request, "❌ حدث خطأ غير متوقع أثناء رفض الطلب. يرجى إعادة المحاولة.")
        return redirect(next_url)


# ------------------------------------------------------------------#
#            Booking Requests List (secretary page)                  #
# ------------------------------------------------------------------#
@secretary_required
@require_GET
def booking_requests_list(request: HttpRequest):
    qs = PatientBookingRequest.objects.all()

    if BookingRequestStatus and _model_has_field(PatientBookingRequest, "status"):
        pending_status = getattr(BookingRequestStatus, "PENDING", None)
        requested_status = getattr(BookingRequestStatus, "REQUESTED", None)

        if pending_status and requested_status:
            qs = qs.filter(status__in=[pending_status, requested_status])
        elif pending_status:
            qs = qs.filter(status=pending_status)
        elif requested_status:
            qs = qs.filter(status=requested_status)

    qs = _filter_booking_requests_for_user(qs, request.user)

    try:
        qs = qs.select_related("doctor", "doctor__user")
    except Exception:
        pass

    qs = (
        qs.order_by("-submitted_at")
        if _model_has_field(PatientBookingRequest, "submitted_at")
        else qs.order_by("-scheduled_time")
    )

    page = Paginator(qs, 20).get_page(request.GET.get("page"))

    if _model_has_field(PatientBookingRequest, "seen_by_secretary") or _model_has_field(PatientBookingRequest, "seen_at"):
        ids = [obj.pk for obj in page.object_list]
        _mark_booking_requests_seen_bulk(ids)

    return render(request, "appointments/booking_requests_list.html", {"requests": page})


# ------------------------------------------------------------------#
#                       Secretary Reports                            #
# ------------------------------------------------------------------#
@secretary_required
@require_GET
def secretary_reports(request: HttpRequest):
    period, start, end = _get_period_range(request)

    base_qs = _filter_by_day_range(
        _active_appts_qs().select_related("patient", "doctor__user"),
        start,
        end,
    )
    base_qs = _filter_appointments_for_user(base_qs, request.user)

    total = base_qs.count()
    completed = base_qs.filter(status=AppointmentStatus.COMPLETED).count()
    cancelled = base_qs.filter(status=AppointmentStatus.CANCELLED).count()
    pending = base_qs.filter(status=AppointmentStatus.PENDING).count()

    called = 0
    if hasattr(AppointmentStatus, "CALLED"):
        called = base_qs.filter(status=getattr(AppointmentStatus, "CALLED")).count()

    revenue = 0
    if _model_has_field(Appointment, "iqd_amount"):
        revenue = base_qs.aggregate(total=Sum("iqd_amount")).get("total") or 0

    if _model_has_field(Appointment, "scheduled_day"):
        daily_qs = base_qs.values("scheduled_day").annotate(count=Count("id")).order_by("scheduled_day")
        daily = [
            {
                "day": row["scheduled_day"].strftime("%Y-%m-%d") if isinstance(row["scheduled_day"], date) else str(row["scheduled_day"]),
                "count": row["count"],
            }
            for row in daily_qs
        ]
    else:
        daily_qs = (
            base_qs.annotate(day=TruncDate("scheduled_time"))
            .values("day")
            .annotate(count=Count("id"))
            .order_by("day")
        )
        daily = [
            {"day": row["day"].strftime("%Y-%m-%d") if isinstance(row["day"], date) else str(row["day"]), "count": row["count"]}
            for row in daily_qs
        ]

    summary = {
        "total": total,
        "completed": completed,
        "cancelled": cancelled,
        "pending": pending,
        "called": called,
        "revenue": revenue,
    }

    ctx = {
        "period": period,
        "start": start,
        "end": end,
        "summary": summary,
        "daily": daily,
        "appointments": base_qs.order_by("scheduled_time"),
    }

    try:
        return render(request, "appointments/secretary_reports.html", ctx)
    except TemplateDoesNotExist:
        return HttpResponse("Secretary reports page is available.", status=200)


@secretary_required
@require_GET
def reports_export(request: HttpRequest):
    fmt = (request.GET.get("format") or "csv").lower()
    _period, start, end = _get_period_range(request)

    qs = _filter_by_day_range(
        _active_appts_qs().select_related("patient", "doctor__user").order_by("scheduled_time"),
        start,
        end,
    )
    qs = _filter_appointments_for_user(qs, request.user)

    headers = ["ID", "Date", "Time", "Doctor", "Patient", "Status", "IQD"]
    rows: list[list[object]] = []

    for a in qs:
        dt = _dt_for_display(a.scheduled_time)
        date_str = dt.strftime("%Y-%m-%d") if dt else ""
        time_str = dt.strftime("%H:%M") if dt else ""
        status_label = a.get_status_display() if hasattr(a, "get_status_display") else str(a.status)
        amount = getattr(a, "iqd_amount", None) or 0
        rows.append([a.id, date_str, time_str, _doctor_name(a.doctor), a.patient.full_name, status_label, amount])

    filename_base = f"clinichub_reports_{start:%Y%m%d}_{end:%Y%m%d}"

    # XLSX (optional)
    if fmt == "xlsx":
        try:
            import openpyxl  # type: ignore
            from openpyxl.utils import get_column_letter  # type: ignore
        except ImportError:
            fmt = "csv"

    if fmt == "xlsx":
        wb = openpyxl.Workbook()  # type: ignore[name-defined]
        ws = wb.active
        ws.title = "Appointments"

        ws.append(headers)
        for row in rows:
            ws.append(row)

        for col_idx, header in enumerate(headers, start=1):
            col_values = [len(str(header))]
            if rows:
                col_values.extend(len(str(r[col_idx - 1])) for r in rows)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(col_values) + 2  # type: ignore[name-defined]

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    # CSV
    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
    resp.write("\ufeff")  # BOM
    writer = csv.writer(resp)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return resp


# ------------------------------------------------------------------#
#                   Secretary Settings & Polling                     #
# ------------------------------------------------------------------#
@secretary_required
@require_http_methods(["GET", "POST"])
def secretary_settings(request: HttpRequest):
    user = request.user

    if request.method == "POST":
        form_type = (request.POST.get("form_type") or "").strip().lower()

        if form_type == "profile":
            profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=user)
            password_form = CustomPasswordForm(user=user)

            if profile_form.is_valid():
                changed = list(profile_form.changed_data or [])
                profile_form.save()

                _audit(
                    request=request,
                    action="update",
                    actor=user,
                    instance=user,
                    message="Secretary profile updated",
                    extra_data={"changed_fields": changed},
                )

                messages.success(
                    request,
                    f"✅ تم تحديث الملف الشخصي ({', '.join(changed)}) بنجاح." if changed else "ℹ لم يتم رصد أي تغييرات.",
                )
                return redirect("appointments:secretary_settings")

            messages.error(request, "⚠️ لم يتم حفظ التعديلات. يرجى تصحيح الأخطاء في نموذج الملف الشخصي.")

        elif form_type == "password":
            profile_form = ProfileUpdateForm(instance=user)
            password_form = CustomPasswordForm(user=user, data=request.POST)

            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, user)
                if request.POST.get("enforce_logout"):
                    _logout_other_sessions(request)

                _audit(
                    request=request,
                    action="update",
                    actor=user,
                    instance=user,
                    message="Secretary password changed",
                    extra_data={"enforce_logout_others": bool(request.POST.get("enforce_logout"))},
                )

                messages.success(request, "🔒 تم تغيير كلمة المرور بنجاح.")
                return redirect("appointments:secretary_settings")

            messages.error(request, "⚠️ لم يتم تغيير كلمة المرور. يرجى التحقق من البيانات المدخلة.")

        else:
            messages.error(request, "⚠️ تم إرسال نموذج غير معروف.")
            profile_form = ProfileUpdateForm(instance=user)
            password_form = CustomPasswordForm(user=user)
    else:
        profile_form = ProfileUpdateForm(instance=user)
        password_form = CustomPasswordForm(user=user)

    return render(
        request,
        "appointments/secretary_settings.html",
        {"profile_form": profile_form, "password_form": password_form},
    )


@secretary_required
@require_GET
def new_booking_requests_api(request: HttpRequest):
    assigned_doctor = _secretary_assigned_doctor(request.user)
    notifs = Notification.objects.filter(is_read=False)

    if not _notif_has_related_request():
        return _json_success({"count": 0, "booking_requests": []})

    notifs = notifs.exclude(related_booking_request__isnull=True)
    if assigned_doctor is not None:
        notifs = notifs.filter(related_booking_request__doctor=assigned_doctor)

    notifs = notifs.select_related("related_booking_request__doctor__user")

    order_field = "-created_at" if _model_has_field(Notification, "created_at") else "-id"
    notifs = notifs.order_by(order_field)

    total_count = notifs.count()
    notifs = notifs[:50]

    items: list[dict] = []
    for n in notifs:
        br = getattr(n, "related_booking_request", None)

        full_name = getattr(br, "full_name", "") if br else ""
        doctor_obj = getattr(br, "doctor", None) if br else None
        doctor_name = _doctor_name(doctor_obj) if doctor_obj else ""
        time_display = _fmt_dt(getattr(br, "scheduled_time", None), "%Y-%m-%d %H:%M") if br else ""
        status = str(getattr(br, "status", "")) if br else ""
        booking_id = getattr(br, "id", None) if br else None

        created_at_val = getattr(n, "created_at", None)
        created_at_str = created_at_val.isoformat() if created_at_val else ""

        items.append(
            {
                "id": booking_id or n.id,
                "full_name": full_name,
                "requested_doctor": doctor_name,
                "requested_time_display": time_display,
                "status": status,
                "title": getattr(n, "title", ""),
                "message": getattr(n, "message", ""),
                "created_at": created_at_str,
                "source": "notification",
            }
        )

    return _json_success({"count": total_count, "booking_requests": items})


@secretary_required
@require_GET
def notifications_list(request: HttpRequest):
    qs = (
        Notification.objects.all().order_by("-created_at", "-pk")
        if _model_has_field(Notification, "created_at")
        else Notification.objects.all().order_by("-pk")
    )

    if _notif_has_related_request():
        assigned = _secretary_assigned_doctor(request.user)
        if assigned is not None:
            qs = qs.filter(Q(related_booking_request__isnull=True) | Q(related_booking_request__doctor=assigned))

    page = Paginator(qs, 30).get_page(request.GET.get("page"))
    return render(request, "appointments/notifications_list.html", {"notifications": page})


@secretary_required
@require_POST
def mark_notification_read(request: HttpRequest, pk: int):
    n = get_object_or_404(Notification, pk=pk)

    if _notif_has_related_request():
        assigned = _secretary_assigned_doctor(request.user)
        if assigned is not None:
            br = getattr(n, "related_booking_request", None)
            if br and getattr(br, "doctor_id", None) != assigned.id:
                return HttpResponseForbidden("Forbidden.")

    Notification.objects.filter(pk=n.pk).update(is_read=True)

    _audit(
        request=request,
        action="update",
        instance=n,
        message="Notification marked as read",
        extra_data={"notification_id": n.pk},
    )

    return _json_success({"id": n.pk, "is_read": True})


@secretary_required
@require_POST
def mark_all_notifications_read(request: HttpRequest):
    qs = Notification.objects.filter(is_read=False)

    if _notif_has_related_request():
        assigned = _secretary_assigned_doctor(request.user)
        if assigned is not None:
            qs = qs.filter(Q(related_booking_request__isnull=True) | Q(related_booking_request__doctor=assigned))

    updated = qs.update(is_read=True)

    _audit(
        request=request,
        action="update",
        actor=request.user,
        message="All visible notifications marked as read",
        extra_data={"updated_count": updated},
    )

    return _json_success({"updated": updated})


# ------------------------------------------------------------------#
#                  Queue Display + APIs                              #
# ------------------------------------------------------------------#
def _queue_doctors_queryset_for(user=None):
    qs = Doctor.objects.select_related("user").order_by("id")
    limit = getattr(settings, "QUEUE_DISPLAY_DOCTORS_LIMIT", None)

    if user is not None:
        assigned = _secretary_assigned_doctor(user)
        if assigned is not None:
            qs = qs.filter(pk=assigned.pk)

    if isinstance(limit, int) and limit > 0:
        qs = qs[:limit]
    return qs


def _queue_public_show_patient_names() -> bool:
    return bool(getattr(settings, "QUEUE_PUBLIC_SHOW_PATIENT_NAME", False))


def _queue_public_include_ids() -> bool:
    return bool(getattr(settings, "QUEUE_PUBLIC_INCLUDE_IDS", False))


def _queue_public_token_required() -> bool:
    token = str(getattr(settings, "QUEUE_PUBLIC_TOKEN", "") or "")
    if token:
        return True
    return bool(getattr(settings, "QUEUE_PUBLIC_REQUIRE_TOKEN", False))


def _queue_public_token_ok(request: HttpRequest) -> bool:
    if not _queue_public_token_required():
        return True

    expected = str(getattr(settings, "QUEUE_PUBLIC_TOKEN", "") or "")
    if not expected:
        return False

    provided = (request.GET.get("token") or "").strip() or (request.headers.get("X-Queue-Token") or "").strip()
    if not provided:
        return False

    return secrets.compare_digest(provided, expected)


def _queue_snapshot_internal(user=None) -> list[dict]:
    today = _today()
    default_mins = int(getattr(settings, "APPOINTMENT_DURATION_MINUTES", 15) or 15)

    appts_qs = (
        _active_appts_qs()
        .filter(status__in=_queue_active_statuses())
        .select_related("patient", "doctor__user")
        .order_by("scheduled_time")
    )
    appts_qs = _filter_by_day(appts_qs, today)
    if user is not None:
        appts_qs = _filter_appointments_for_user(appts_qs, user)

    appts = list(appts_qs)
    doctors = list(_queue_doctors_queryset_for(user=user))

    by_doc: dict[int, list[Appointment]] = {}
    for a in appts:
        by_doc.setdefault(a.doctor_id, []).append(a)

    queues: list[dict] = []
    for d in doctors:
        today_appts = by_doc.get(d.id, [])

        current_obj = None
        if hasattr(AppointmentStatus, "CALLED"):
            called_status = getattr(AppointmentStatus, "CALLED")
            for a in today_appts:
                if a.status == called_status:
                    current_obj = a
                    break
        if current_obj is None and today_appts:
            current_obj = today_appts[0]

        waiting_objs = [a for a in today_appts if (current_obj is None or a.id != current_obj.id)]

        current = None
        waiting: list[dict[str, Any]] = []

        if current_obj:
            current = {
                "id": current_obj.id,
                "number": _format_queue_number(getattr(current_obj, "queue_number", None)),
                "patient_name": current_obj.patient.full_name,
                "time": _fmt_dt(current_obj.scheduled_time, "%H:%M"),
                "status": str(current_obj.status),
            }

        for w in waiting_objs:
            waiting.append(
                {
                    "id": w.id,
                    "number": _format_queue_number(getattr(w, "queue_number", None)),
                    "patient_name": w.patient.full_name,
                    "time": _fmt_dt(w.scheduled_time, "%H:%M"),
                    "status": str(w.status),
                }
            )

        next_queue = current["number"] if current else "No appointments"
        queues.append(
            {
                "doctor_id": d.id,
                "doctor_name": _doctor_name(d),
                "status": "available" if today_appts else "on_break",
                "next_queue": next_queue,
                "waiting_count": len(waiting),
                "current": current,
                "waiting": waiting,
                "avg_time": default_mins,
            }
        )

    return queues


def _queue_snapshot_public() -> list[dict]:
    today = _today()
    default_mins = int(getattr(settings, "APPOINTMENT_DURATION_MINUTES", 15) or 15)

    show_names = _queue_public_show_patient_names()
    include_ids = _queue_public_include_ids()

    appts_qs = (
        _active_appts_qs()
        .filter(status__in=_queue_active_statuses())
        .select_related("patient", "doctor__user")
        .order_by("scheduled_time")
    )
    appts = list(_filter_by_day(appts_qs, today))
    doctors = list(_queue_doctors_queryset_for(user=None))

    by_doc: dict[int, list[Appointment]] = {}
    for a in appts:
        by_doc.setdefault(a.doctor_id, []).append(a)

    queues: list[dict] = []
    for d in doctors:
        today_appts = by_doc.get(d.id, [])

        current_obj = None
        if hasattr(AppointmentStatus, "CALLED"):
            called_status = getattr(AppointmentStatus, "CALLED")
            for a in today_appts:
                if a.status == called_status:
                    current_obj = a
                    break
        if current_obj is None and today_appts:
            current_obj = today_appts[0]

        waiting_objs = [a for a in today_appts if (current_obj is None or a.id != current_obj.id)]

        current: dict[str, Any] | None = None
        waiting: list[dict[str, Any]] = []

        if current_obj:
            current = {
                "number": _format_queue_number(getattr(current_obj, "queue_number", None)),
                "time": _fmt_dt(current_obj.scheduled_time, "%H:%M"),
                "status": str(current_obj.status),
            }
            if include_ids:
                current["id"] = current_obj.id
            if show_names:
                current["patient_name"] = current_obj.patient.full_name

        for w in waiting_objs:
            item: dict[str, Any] = {
                "number": _format_queue_number(getattr(w, "queue_number", None)),
                "time": _fmt_dt(w.scheduled_time, "%H:%M"),
                "status": str(w.status),
            }
            if include_ids:
                item["id"] = w.id
            if show_names:
                item["patient_name"] = w.patient.full_name
            waiting.append(item)

        next_queue = current["number"] if current else "No appointments"
        queues.append(
            {
                "doctor_id": d.id,
                "doctor_name": _doctor_name(d),
                "status": "available" if today_appts else "on_break",
                "next_queue": next_queue,
                "waiting_count": len(waiting),
                "current": current,
                "waiting": waiting,
                "avg_time": default_mins,
            }
        )

    return queues


@require_GET
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def queue_display(request: HttpRequest):
    if _queue_public_token_required() and not _queue_public_token_ok(request):
        return HttpResponseForbidden("Queue display is protected.")
    try:
        return render(request, "appointments/queue_display.html", {"queues": _queue_snapshot_public()})
    except TemplateDoesNotExist:
        return HttpResponse("Queue display is available.", status=200)


@require_GET
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def queue_public_api(request: HttpRequest):
    if _queue_public_token_required() and not _queue_public_token_ok(request):
        return _json_error("Forbidden", status=403)
    return _json_success({"queues": _queue_snapshot_public()})


@require_GET
@cache_control(no_cache=True, no_store=True, must_revalidate=True)
def queue_number_api(request: HttpRequest):
    user = getattr(request, "user", None)
    is_staff_view = bool(
        user and user.is_authenticated and (_user_is_secretary(user) or getattr(user, "is_superuser", False))
    )

    if is_staff_view:
        return _json_success({"queues": _queue_snapshot_internal(user)})

    if _queue_public_token_required() and not _queue_public_token_ok(request):
        return _json_error("Forbidden", status=403)

    return _json_success({"queues": _queue_snapshot_public()})


@secretary_required
@require_POST
def call_next_api(request: HttpRequest, doctor_id: int):
    today = _today()

    assigned_doctor = _secretary_assigned_doctor(request.user)
    if assigned_doctor is not None and assigned_doctor.id != doctor_id:
        return _json_error("You cannot control the queue of another doctor.", status=403)

    # Accept POST form OR JSON body
    appt_id_raw = ""
    if request.content_type and "application/json" in (request.content_type or ""):
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
            appt_id_raw = str(payload.get("appointment_id") or "").strip()
        except Exception:
            appt_id_raw = ""
    else:
        appt_id_raw = (request.POST.get("appointment_id") or "").strip()

    appt_id: int | None = None
    if appt_id_raw:
        try:
            appt_id = int(appt_id_raw)
        except Exception:
            appt_id = None

    use_called_mode = _queue_use_called_status()
    called_status = getattr(AppointmentStatus, "CALLED", None)
    completed_status = AppointmentStatus.COMPLETED
    waiting_statuses = _queue_waiting_statuses()
    active_statuses = _queue_active_statuses()

    with transaction.atomic():
        base = (
            _active_appts_qs()
            .select_for_update()
            .filter(doctor_id=doctor_id)
            .select_related("patient", "doctor__user")
            .order_by("scheduled_time")
        )
        base = _filter_by_day(base, today)

        if use_called_mode and called_status is not None:
            current_called = base.filter(status=called_status).first()
            if current_called:
                Appointment.objects.filter(pk=current_called.pk).update(status=completed_status)

            next_qs = base.filter(status__in=waiting_statuses)
            nxt = next_qs.filter(pk=appt_id).first() if appt_id else None
            if not nxt:
                nxt = next_qs.first()
            if not nxt:
                return _json_error("No waiting appointments for this doctor.", status=404)

            Appointment.objects.filter(pk=nxt.pk).update(status=called_status)
            return _json_success({"queues": _queue_snapshot_internal(request.user)})

        # Complete-mode: mark next active as completed
        next_qs = base.filter(status__in=active_statuses).exclude(status=AppointmentStatus.CANCELLED)
        next_qs = next_qs.exclude(status=completed_status)

        nxt = next_qs.filter(pk=appt_id).first() if appt_id else None
        if not nxt:
            nxt = next_qs.first()
        if not nxt:
            return _json_error("No waiting appointments for this doctor.", status=404)

        Appointment.objects.filter(pk=nxt.pk).update(status=completed_status)
        return _json_success({"queues": _queue_snapshot_internal(request.user)})


@secretary_required
@require_GET
def current_patient_api(request: HttpRequest):
    today = _today()

    pend_qs = (
        _active_appts_qs()
        .filter(status__in=_queue_active_statuses())
        .order_by("scheduled_time")
        .select_related("patient", "doctor__user")
    )
    pend_qs = _filter_by_day(pend_qs, today)
    pend_qs = _filter_appointments_for_user(pend_qs, request.user)

    pend = list(pend_qs[:20])

    current_obj = None
    if hasattr(AppointmentStatus, "CALLED"):
        called_status = getattr(AppointmentStatus, "CALLED")
        for a in pend:
            if a.status == called_status:
                current_obj = a
                break
    if current_obj is None and pend:
        current_obj = pend[0]

    current = None
    if current_obj:
        current = {
            "id": current_obj.id,
            "number": _format_queue_number(getattr(current_obj, "queue_number", None)),
            "patient_name": current_obj.patient.full_name,
            "doctor_name": _doctor_name(current_obj.doctor),
            "status": str(current_obj.status),
        }

    return _json_success({"current_patient": current})